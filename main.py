from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
import csv
import os
from datetime import datetime
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


@app.get("/", response_class=HTMLResponse)
def intro(request: Request):
    return templates.TemplateResponse(
        "intro.html",
        {
            "request": request,
            "names": "Рустам & Дарья"
        }
    )


@app.get("/invite", response_class=HTMLResponse)
def invite(request: Request):
    deadline = datetime(2026, 4, 1)
    now = datetime.now()

    rsvp_open = now < deadline

    return templates.TemplateResponse(
        "invite.html",
        {
            "request": request,
            "rsvp_open": rsvp_open
        }
    )


@app.get("/rsvp", response_class=HTMLResponse)
def rsvp_form(request: Request):
    return templates.TemplateResponse("rsvp.html", {"request": request})


from openpyxl import Workbook, load_workbook
from pathlib import Path

@app.post("/rsvp")
def rsvp_submit(
    name: str = Form(...),
    attending: str = Form(...),
    guests_count: int = Form(0),
    with_children: str = Form("no"),
    children_count: int = Form(0),
    transfer: str = Form("no"),
    alcohol: list[str] = Form([]),
    comment: str = Form("")
):
    os.makedirs("data", exist_ok=True)
    file_path = Path("data/responses.xlsx")

    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Дата",
            "Имя",
            "Присутствие",
            "Кол-во гостей",
            "С детьми",
            "Кол-во детей",
            "Трансфер",
            "Алкоголь",
            "Комментарий"
        ])

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        name,
        attending,
        guests_count,
        with_children,
        children_count,
        transfer,
        ", ".join(alcohol),
        comment
    ])

    wb.save(file_path)

    return RedirectResponse("/thanks", status_code=303)



@app.get("/thanks", response_class=HTMLResponse)
def thanks(request: Request):
    return templates.TemplateResponse("thanks.html", {"request": request})
